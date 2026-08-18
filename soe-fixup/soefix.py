#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl", "pyperclip"]
# ///
"""soefix - bake per-store SOE payloads onto the clipboard.

Everything is derived from the site number N: RHS02 = 10.56.N.93, VM SOE = 10.56.N.1.

Paste on RHS02 (the PSM session):
  soefix restore N                     copy X:\\SOE_Backup, run the Server2022 restore, list drivers
  soefix push N --driver "<folder>"    stage convert.ps1 + Maxtel + driver + generatekvs/JRE onto the SOE
                                       (<folder> = name under X:\\Certeq, e.g. "Printer Drivers\\Epson", or a full X:\\ path)
  soefix push N --cleanup              stage cleanup.ps1 (+ generatekvs) on an already-converted SOE
  soefix verify N                      post-reboot checks over c$, pull the summary back to the Mac
  soefix tidy N                        last step: remove soefix's own files from the SOE (scripts, logs, .bak)

Type on the SOE (Win+R):   C:\\Temp\\soefix\\go

ac:
  soefix log N                         ingest SOE_Static_Files/soefix-logs/N.txt into results.log
  soefix log N <note>                  manual entry;  soefix log  (no args) = clipboard mode
  soefix list                          pending stores from the sheet with corrected day counts

Options: --name "Store Name" (when N is not in the sheet; remembered per site), --force (ignore Done in the sheet),
         --ip 10.56.55.1 (SOE address when the third octet is not the site number, e.g. sites > 255;
         RHS02 is taken as .93 on the same subnet; remembered per site in sites.json).
Paths (sheet, static folder, its tsclient UNC, results log) come from soefix.toml next to
this file - see soefix.example.toml. Works on macOS and Windows (uv + pyperclip).

Day counts come from the sheet's 'Days Since Completion' column, an XLOOKUP into an
external workbook - Excel only stores the value cached at last save, so we add the days
elapsed since the file was modified.
"""

import json
import re
import sys
import tomllib
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pyperclip

HERE = Path(__file__).resolve().parent
TEMPLATES = HERE / "templates"
CONFIG = HERE / "soefix.toml"
HEADER_ROW = 4  # Ref ID | Site # | Site Name | Allocated | Completed | Notes | Days Since Completion


def _load_config() -> dict:
    if not CONFIG.exists():
        return {}
    with CONFIG.open("rb") as f:
        return tomllib.load(f)


def _path(cfg: dict, key: str) -> Path | None:
    return Path(cfg[key]).expanduser() if cfg.get(key) else None


_cfg = _load_config()
SHEET = _path(_cfg, "sheet")                       # tracking workbook (read-only)
STATIC = _path(_cfg, "static_dir")                 # local generatekvs.exe / JRE folder
LOGS_IN = STATIC / "soefix-logs" if STATIC else None
STATIC_UNC = _cfg.get("static_unc", r"\\tsclient\SOE_Static_Files")  # same folder as seen from RHS02
LOG = _path(_cfg, "results_log") or HERE / "results.log"


def die(msg: str) -> None:
    print(f"soefix: {msg}", file=sys.stderr)
    sys.exit(1)


# --------------------------------------------------------------------------- sheet
def load_stores() -> tuple[list[dict], int]:
    """Return (stores, staleness_days). Day counts are already staleness-corrected."""
    if SHEET is None:
        die(f"no 'sheet' configured - create {CONFIG.name} next to soefix.py (see soefix.example.toml)")
    if SHEET is None:
        die(f"no 'sheet' configured - create {CONFIG.name} next to soefix.py (see soefix.example.toml)")
    if not SHEET.exists():
        die(f"spreadsheet not found: {SHEET}")
    staleness = (date.today() - datetime.fromtimestamp(SHEET.stat().st_mtime).date()).days
    wb = openpyxl.load_workbook(SHEET, data_only=True, read_only=True)
    ws = wb.worksheets[0]  # the tracking sheet (was 'Sheet1', now 'Track') - always first
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[:7]]
    if header[1] != "Site #" or not header[6].startswith("Days Since"):
        die(f"unexpected header row {HEADER_ROW} in {SHEET.name}: {header}")
    stores = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        ref, site, name, _alloc, completed, _notes, days = row[:7]
        checks = str(row[8] or "") if len(row) > 8 else ""
        if site is None:
            continue
        stores.append(
            {
                "ref": ref,
                "site": int(site),
                "name": str(name).strip(),
                "done": str(completed or "").strip().lower() == "done",
                "days": days + staleness if isinstance(days, (int, float)) else days,
                # column I from the 2026-08-13 run: "recollect NOT run" = generatekvs was 2015
                "backlog": "recollect NOT run" in checks,
            }
        )
    wb.close()
    return stores, staleness


def find_store(stores: list[dict], query: str) -> dict:
    if query.isdigit():
        matches = [s for s in stores if s["site"] == int(query)]
    else:
        matches = [s for s in stores if query.lower() in s["name"].lower()]
    if not matches:
        die(f"no store matching {query!r} in the sheet")
    if len(matches) > 1:
        names = ", ".join(f"{s['site']} {s['name']}" for s in matches)
        die(f"{query!r} is ambiguous: {names}")
    return matches[0]


def store_info(site: int, name: str | None) -> dict:
    """Sheet row for the site if present, else a bare fallback. --name always wins."""
    stores, _ = load_stores()
    row = next((s for s in stores if s["site"] == site), None)
    if row:
        info = {"site": site, "name": row["name"], "ref": row["ref"] or "", "days": row["days"],
                "done": row["done"], "backlog": row.get("backlog", False)}
    else:
        info = {"site": site, "name": _site_rec(site).get("name") or f"site {site}", "ref": "", "days": None,
                "done": False, "backlog": False}
    if name:
        info["name"] = name
        remember(site, name=name)   # sticky, like --ip
    return info


# --------------------------------------------------------------------------- rendering
IP_OVERRIDE: str | None = None  # --ip 10.56.55.1 (sites whose third octet is not the site number)
SITES = HERE / "sites.json"     # remembered --ip per site (local, git-ignored)


def _sites() -> dict:
    if SITES.exists():
        try:
            return json.loads(SITES.read_text())
        except ValueError:
            return {}
    return {}


def _site_rec(site: int) -> dict:
    rec = _sites().get(str(site), {})
    return {"ip": rec} if isinstance(rec, str) else rec   # older files stored just the ip


def remember(site: int, **kv: str) -> None:
    d = _sites()
    rec = _site_rec(site)
    rec.update({k: v for k, v in kv.items() if v})
    if d.get(str(site)) != rec:
        d[str(site)] = rec
        SITES.write_text(json.dumps(d, indent=2, sort_keys=True) + "\n")


def driver_leaf(driver: str) -> str:
    return driver.replace("/", "\\").rstrip("\\").split("\\")[-1] if driver else ""


def derive(site: int) -> dict:
    ip = IP_OVERRIDE or _site_rec(site).get("ip")
    if ip:
        parts = ip.split(".")
        if len(parts) != 4 or not all(p.isdigit() and 0 <= int(p) <= 255 for p in parts):
            die(f"--ip must be a full IPv4 address like 10.56.55.1, got {ip!r}")
        if IP_OVERRIDE:
            remember(site, ip=ip)
        net = ".".join(parts[:3])
    else:
        if site > 255:
            die(f"site {site} can't be the third octet of the IP - pass the SOE address once, e.g. --ip 10.56.55.1 (remembered in {SITES.name})")
        net = f"10.56.{site}"
    return {"site": site, "ip_rhs": f"{net}.93", "ip_soe": f"{net}.1"}


def ps_quote(s: str) -> str:
    """Escape for a PowerShell single-quoted string."""
    return s.replace("'", "''")


def render(template: str, **vars: str) -> str:
    text = (TEMPLATES / template).read_text()
    if "{{CONNECT}}" in text:
        vars = {"CONNECT": (TEMPLATES / "_connect.ps1").read_text().rstrip("\n"), **vars}
    for k, v in vars.items():
        text = text.replace("{{" + k + "}}", str(v))
    left = re.findall(r"\{\{[A-Z_]+\}\}", text)
    if left:
        raise ValueError(f"{template}: unfilled placeholders {sorted(set(left))}")
    return text


def bake_restore(site: int) -> str:
    d = derive(site)
    return render("restore.ps1", SITE=site, IP_SOE=d["ip_soe"])


def bake_verify(site: int, name: str) -> str:
    d = derive(site)
    return render("verify.ps1", SITE=site, NAME=ps_quote(name), IP_SOE=d["ip_soe"], STATIC_UNC=ps_quote(STATIC_UNC))


def bake_soe(mode: str, info: dict, driver: str, ip_rhs: str) -> str:
    """Render the script that runs ON the SOE (convert.ps1 or cleanup.ps1)."""
    common = dict(SITE=info["site"], NAME=ps_quote(info["name"]), REF=ps_quote(str(info["ref"])))
    if mode == "convert":
        # the SOE only ever sees C:\Temp\<leaf>, whatever form --driver took
        return render("convert.ps1", DRIVER=ps_quote(driver_leaf(driver)), IP_RHS=ip_rhs, **common)
    if mode == "cleanup":
        if not isinstance(info["days"], (int, float)):
            raise ValueError("cleanup mode needs a day count from the sheet (Days Since Completion is empty)")
        return render("cleanup.ps1", DAYS=int(info["days"]), **common)
    raise ValueError(f"unknown mode {mode!r}")


def check_embeddable(script: str) -> None:
    for i, line in enumerate(script.splitlines(), 1):
        if line.startswith("'@"):
            raise ValueError(f"embedded script line {i} starts with '@ - would terminate the here-string")


def bake_push(site: int, driver: str, cleanup: bool, name: str | None, force: bool) -> str:
    info = store_info(site, name)
    d = derive(site)
    mode = "cleanup" if cleanup else "convert"
    if info["done"] and not force and not (cleanup and info["backlog"]):
        die(f"site {site} {info['name']} is marked Done in the sheet - use --force to push anyway")
    soe_script = bake_soe(mode, info, driver, d["ip_rhs"])
    check_embeddable(soe_script)
    if driver:
        remember(site, driver=driver_leaf(driver))
    script_name = f"{mode}.ps1"
    go = render("go.cmd", SCRIPT=script_name)
    return render(
        "push.ps1",
        SITE=site,
        NAME=ps_quote(info["name"]),
        MODE=mode,
        DRIVER=ps_quote(driver),
        IP_SOE=d["ip_soe"],
        STATIC_UNC=ps_quote(STATIC_UNC),
        SCRIPT_NAME=script_name,
        SOE_SCRIPT=soe_script,
        GO_CMD=go,
    )


def bake_tidy(site: int, name: str) -> str:
    d = derive(site)
    return render("tidy.ps1", SITE=site, NAME=ps_quote(name), IP_SOE=d["ip_soe"], STATIC_UNC=ps_quote(STATIC_UNC))


# --------------------------------------------------------------------------- clipboard
def pbcopy(text: str) -> None:
    pyperclip.copy(text)


def pbpaste() -> str:
    return pyperclip.paste()


# --------------------------------------------------------------------------- commands
def cmd_restore(site: int) -> None:
    pbcopy(bake_restore(site))
    print(f"restore payload for site {site} on clipboard (RHS02 = {derive(site)['ip_rhs']}).")
    print("Paste into a normal PowerShell on RHS02 (X: must be visible). It copies X:\\SOE_Backup, runs the restore")
    print("(~10 min, don't interrupt), then lists the X:\\Certeq driver folders for `soefix push --driver`.")


def cmd_push(site: int, driver: str, cleanup: bool, name: str | None, force: bool) -> None:
    payload = bake_push(site, driver, cleanup, name, force)
    pbcopy(payload)
    info = store_info(site, name)
    d = derive(site)
    mode = "cleanup" if cleanup else "convert"
    print(f"push payload ({mode}) for site {site} {info['name']} on clipboard - SOE = {d['ip_soe']} ({len(payload) // 1024} KB).")
    if mode == "convert" and not driver:
        print("note: no --driver given - the SOE script will pause for you to copy the driver folder and add it by hand.")
    if mode == "cleanup":
        why = " (sheet: recollect not run on the 13/08 pass)" if info["backlog"] else ""
        print(f"cleanup will recollect with generatekvs.exe /auto {int(info['days'])}{why}.")
    print("Paste into PowerShell on RHS02 (VM must be up, you logged in as Administrator on it).")
    print("Then on the SOE:  Win+R  ->  C:\\Temp\\soefix\\go")


def cmd_verify(site: int, name: str | None) -> None:
    info = store_info(site, name)
    pbcopy(bake_verify(site, info["name"]))
    print(f"verify payload for site {site} {info['name']} on clipboard - paste on RHS02 after the SOE reboot.")
    print(f"It also copies the SOE summary to {STATIC_UNC}\\soefix-logs\\{site}.txt (= {LOGS_IN}) - then: soefix log {site}")


def cmd_tidy(site: int, name: str | None) -> None:
    info = store_info(site, name)
    pbcopy(bake_tidy(site, info["name"]))
    print(f"tidy payload for site {site} {info['name']} on clipboard - paste on RHS02 as the last step.")
    print("Removes from the SOE: C:\\Temp\\soefix (scripts, summary, transcript), C:\\Helpdesk\\soe_fixup_summary.txt,")
    print("generatekvs.exe.2015.bak. Driver folder, Maxtel, JRE and RHS02's C:\\SOE_Backup are left in place.")


def cmd_list() -> None:
    stores, staleness = load_stores()
    pending = [s for s in stores if not s["done"]]
    if staleness:
        print(f"(sheet last saved {staleness}d ago - day counts corrected by +{staleness})\n")
    for s in sorted(pending, key=lambda s: s["site"]):
        print(f"  {s['site']:>4}  {s['name']:<24} {s['days']} days  (ref {s['ref']})")
    print(f"\n{len(pending)} pending, {len(stores) - len(pending)} done")
    backlog = [s for s in stores if s["done"] and s["backlog"]]
    if backlog:
        print(f"\nrecollect backlog (2015 generatekvs on the 13/08 pass) - `soefix push N --cleanup` works without --force:")
        for s in sorted(backlog, key=lambda s: s["site"]):
            print(f"  {s['site']:>4}  {s['name']:<24} /auto {s['days']}  (ref {s['ref']})")


def _append_log(text: str) -> None:
    existing = LOG.read_text() if LOG.exists() else ""
    first = text.splitlines()[0] if text.strip() else ""
    if first and first in existing:
        print(f"note: this entry is already in {LOG.name}; appending again anyway")
    with LOG.open("a") as f:
        f.write(text.rstrip("\n") + "\n\n")
    fails = [l for l in text.splitlines() if "FAIL" in l]
    if fails:
        print("! this store has FAIL lines - check them:")
        for l in fails:
            print(f"  {l}")


def cmd_log(args: list[str]) -> None:
    if len(args) == 1 and args[0].isdigit() and LOGS_IN and (LOGS_IN / f"{args[0]}.txt").exists():
        # file mode: summary pulled back by `soefix verify`
        src = LOGS_IN / f"{args[0]}.txt"
        _append_log(src.read_text(errors="replace").strip())
        print(f"logged site {args[0]} from {src}")
    elif args:
        # manual mode: soefix log <site# | name> [free-text note]
        stores, _ = load_stores()
        s = find_store(stores, args[0])
        note = " ".join(args[1:]) or "OK"
        days = int(s["days"]) if isinstance(s["days"], (int, float)) else "?"
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        tag = f"site {s['site']} {s['name']} (ref {s['ref']})"
        _append_log(f"{stamp}  {tag}  /auto {days}  - {note}")
        print(f"logged: {tag} - {note}")
    else:
        # clipboard mode: summary pushed by the payload via `clip` in-session
        text = pbpaste().strip()
        if not text.startswith("==== SOE FIXUP SUMMARY"):
            die(
                "clipboard doesn't contain a fixup summary - after `soefix verify N` use\n"
                "`soefix log N` (file mode), or log manually: soefix log <site#> [note]"
            )
        _append_log(text)
    entries = LOG.read_text().count("\n\n") if LOG.exists() else 0
    print(f"({entries} entr{'y' if entries == 1 else 'ies'} in {LOG})")


def main() -> None:
    argv = sys.argv[1:]
    force = "--force" in argv
    cleanup = "--cleanup" in argv
    driver = ""
    name = None
    rest = []
    global IP_OVERRIDE
    it = iter(argv)
    for a in it:
        if a in ("--force", "--cleanup"):
            continue
        if a == "--driver":
            driver = next(it, "")
        elif a.startswith("--driver="):
            driver = a.split("=", 1)[1]
        elif a == "--name":
            name = next(it, None)
        elif a.startswith("--name="):
            name = a.split("=", 1)[1]
        elif a == "--ip":
            IP_OVERRIDE = next(it, None)
        elif a.startswith("--ip="):
            IP_OVERRIDE = a.split("=", 1)[1]
        else:
            rest.append(a)
    if not rest:
        print(__doc__.strip())
        sys.exit(0)
    cmd, args = rest[0], rest[1:]

    def site_arg() -> int:
        if not args or not args[0].isdigit():
            die(f"usage: soefix {cmd} <site#>")
        return int(args[0])

    if cmd == "list":
        cmd_list()
    elif cmd == "log":
        cmd_log(args)
    elif cmd == "restore":
        cmd_restore(site_arg())
    elif cmd == "push":
        cmd_push(site_arg(), driver, cleanup, name, force)
    elif cmd == "verify":
        cmd_verify(site_arg(), name)
    elif cmd == "tidy":
        cmd_tidy(site_arg(), name)
    elif cmd.isdigit():
        die(f"the one-paste cleanup is gone - use:  soefix push {cmd} --cleanup   (or restore/push/verify for a conversion)")
    else:
        die(f"unknown command {cmd!r}\n\n{__doc__.strip()}")


if __name__ == "__main__":
    main()
